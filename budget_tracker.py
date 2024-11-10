"""
Budget Tracker Module
"""

import os
import csv
from datetime import datetime
import tkinter as tk
from tkinter import ttk, simpledialog, messagebox
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np
import pdb  # Importing pdb for debugging

# Global variables for workbook and worksheet
WB = None
WS = None

def setup_excel():
    """Creates and/or loads the Excel workbook."""
    global WB, WS
    if not os.path.exists("input_data.xlsx"):
        WB = Workbook()
        WS = WB.active
        WS.title = "Data Input"
        WS.append(["Description", "Amount", "Category", "Date"])
        WB.save("input_data.xlsx")
    else:
        WB = load_workbook("input_data.xlsx")
        WS = WB.active

setup_excel()

def save_to_excel():
    """Saves input data to the Excel file."""
    category = category_combo.get()
    if category.lower() == "income":
        description = income_description_entry.get()
        amount = income_amount_entry.get()
    else:
        description = description_entry.get()
        amount = amount_entry.get()

    date_str = date_entry.get()

    if not description or not amount or not category or not date_str:
        status_label.config(text="Please enter description, amount, category, and date.")
        return

    try:
        amount_value = float(amount)
        date_time = datetime.strptime(date_str, "%m/%d/%Y")
        if category.lower() != "income":
            amount_value = -amount_value  # Expenses are negative

        WS.append([description, amount_value, category, date_str])
        WB.save("input_data.xlsx")
        description_entry.delete(0, tk.END)
        amount_entry.delete(0, tk.END)
        income_description_entry.delete(0, tk.END)
        income_amount_entry.delete(0, tk.END)
        category_combo.set("")
        date_entry.delete(0, tk.END)
        date_entry.insert(0, datetime.now().strftime("%m/%d/%Y"))
        status_label.config(text="Data saved!")
        update_gui()
    except ValueError:
        status_label.config(text="Invalid amount or date. Please enter valid values.")

def read_from_excel():
    """Reads data from the Excel file and updates the Treeview."""
    for item in tree.get_children():
        tree.delete(item)
    for i, row in enumerate(WS.iter_rows(values_only=True)):
        if i == 0:  # Skip header row
            continue
        formatted_row = (
            row[0],  # Description
            f"{row[1]:.2f}",  # Amount formatted to 2 decimal places
            row[2],  # Category
            row[3]   # Date
        )
        tree.insert("", "end", iid=i, values=formatted_row, tags=('evenrow' if i % 2 == 0 else 'oddrow'))
    calculate_total()

def calculate_total():
    """Calculates total expenses and income balance."""
    total_spent = sum(float(row[1]) for row in WS.iter_rows(values_only=True)
                      if row[1] and row[1] != "Amount" and float(row[1]) < 0)
    total_income = sum(float(row[1]) for row in WS.iter_rows(values_only=True)
                       if row[1] and row[1] != "Amount" and float(row[1]) > 0)
    
    total_label.config(text=f"Total Expenses: ${-total_spent:.2f}", font=("Helvetica", 12, "bold"))
    income_label_display.config(text=f"Total Income: ${total_income:.2f}", font=("Helvetica", 12, "bold"))
    balance = total_income + total_spent
    balance_label.config(text=f"Balance: ${balance:.2f}", font=("Helvetica", 12, "bold"))

def update_gui():
    """Updates the GUI by reading data from Excel and recalculating totals."""
    read_from_excel()
    calculate_total()
    update_charts_window()

def export_to_csv():
    """Exports data to CSV for a specific month and year."""
    dialog = simpledialog.askstring("Input", "Enter month and year (MM-YYYY):")
    if not dialog or "-" not in dialog:
        status_label.config(text="Invalid input. Please enter in MM-YYYY format.")
        return
    month, year = map(int, dialog.split('-'))
    if not 1 <= month <= 12 or not 2000 <= year <= 2100:
        status_label.config(text="Invalid month or year.")
        return

    with open(f'budget_data_{month}_{year}.csv', 'w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerow(["Description", "Amount", "Category", "Date"])  # Add header
        for i, row in enumerate(WS.iter_rows(values_only=True)):
            if i == 0:  # Skip the header row
                continue
            date_str = row[3]
            row_date = datetime.strptime(date_str, "%m/%d/%Y")
            if row_date.month == month and row_date.year == year:
                writer.writerow(row)
    status_label.config(text=f"Data for {month:02}/{year} exported to CSV!")

def open_charts_window():
    """Opens a new window to display pie and line charts."""
    global chart_window, canvas_pie, canvas_line, ax_pie, ax_line

    chart_window = tk.Toplevel(root)
    chart_window.title("Charts Window")
    chart_window.geometry("800x600")
    chart_window.configure(bg='#2E2E2E')  # Dark grey background

    figure_pie = plt.Figure(figsize=(5, 4), dpi=100)
    ax_pie = figure_pie.add_subplot(111)
    figure_pie.patch.set_facecolor('#2E2E2E')
    ax_pie.set_facecolor('#2E2E2E')
    canvas_pie = FigureCanvasTkAgg(figure_pie, chart_window)
    canvas_pie.get_tk_widget().pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    figure_line = plt.Figure(figsize=(5, 4), dpi=100)
    ax_line = figure_line.add_subplot(111)
    figure_line.patch.set_facecolor('#2E2E2E')
    ax_line.set_facecolor('#2E2E2E')
    canvas_line = FigureCanvasTkAgg(figure_line, chart_window)
    canvas_line.get_tk_widget().pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    update_charts(ax_pie, ax_line, canvas_pie, canvas_line)

def update_charts_window():
    """Updates the charts in the charts window."""
    if 'ax_pie' in globals() and 'ax_line' in globals():
        update_charts(ax_pie, ax_line, canvas_pie, canvas_line)

def update_charts(ax_pie, ax_line, canvas_pie, canvas_line):
    """Updates the pie and line charts with the latest data."""
    categories = {}
    monthly_data = {}

    abbreviated_categories = {
        "Groceries": "Gro.",
        "Utilities": "Util.",
        "Rent/Mortgage": "Rent",
        "Entertainment": "Ent.",
        "Transportation": "Trans.",
        "Other": "Other"
    }

    for row in WS.iter_rows(values_only=True):
        if row[2] and row[1] and isinstance(row[1], (int, float)):
            category = abbreviated_categories.get(row[2], row[2])
            amount = float(row[1])
            date_str = row[3]
            row_date = datetime.strptime(date_str, "%m/%d/%Y")
            month_year = row_date.strftime("%Y-%m")

            if amount < 0:
                if category in categories:
                    categories[category] += amount
                else:
                    categories[category] = amount

            if month_year not in monthly_data:
                monthly_data[month_year] = {'balance': 0}
            monthly_data[month_year]['balance'] += amount

    ax_pie.clear()
    ax_line.clear()

    if ax_line.get_legend():
        ax_line.get_legend().remove()

    if categories:
        labels = list(categories.keys())
        sizes = [-value for value in categories.values()]

        def reposition_autotexts(wedges, texts, autotexts):
            for i, autotext in enumerate(autotexts):
                angle = (wedges[i].theta2 - wedges[i].theta1) / 2 + wedges[i].theta1
                x = wedges[i].r * np.cos(np.radians(angle))
                y = wedges[i].r * np.sin(np.radians(angle))
                x_offset = x * 0.9
                y_offset = y * 0.9
                autotext.set_position((x_offset + 0.02, y_offset + 0.05))

        def autopct_format(pct):
            return f'{pct:.2f}%'

        wedges, texts, autotexts = ax_pie.pie(
            sizes, labels=labels, autopct=autopct_format, startangle=140,
            colors=['#FF4500', '#FFA500', '#6A5ACD', '#20B2AA', '#FFD700', '#FF69B4'],
            textprops={'color': 'white'}
        )

        reposition_autotexts(wedges, texts, autotexts)

        for text in texts:
            text.set_color('white')
        for autotext in autotexts:
            autotext.set_color('white')
        for wedge in wedges:
            wedge.set_edgecolor('black')
            wedge.set_linewidth(.5)

        ax_pie.axis('equal')
        ax_pie.set_title('Spending by Category', color='white')

    dates = sorted(monthly_data.keys())
    balances = [monthly_data[date]['balance'] for date in dates]

    ax_line.plot(dates, balances, label='Balance', color='white', marker='o')

    ax_line.set_title('Monthly Balance', color='white')
    ax_line.set_xlabel('Month-Year', color='white')
    ax_line.set_ylabel('Balance', color='white')

    legend = ax_line.legend(loc='upper left', frameon=False)
    for text in legend.get_texts():
        text.set_color('white')

    for label in ax_line.get_xticklabels():
        label.set_color('white')
    for label in ax_line.get_yticklabels():
        label.set_color('white')

    ax_line.spines['bottom'].set_color('white')
    ax_line.spines['top'].set_color('white')
    ax_line.spines['left'].set_color('white')
    ax_line.spines['right'].set_color('white')
    ax_line.tick_params(axis='x', colors='white')
    ax_line.tick_params(axis='y', colors='white')

    canvas_pie.draw()
    canvas_line.draw()

def on_treeview_double_click(event):
    """Handles double-click event to edit Treeview entries."""
    item_id = tree.selection()[0]
    item_values = tree.item(item_id, 'values')
    column_index = tree.identify_column(event.x)[1:]  # Get column index
    column_index = int(column_index) - 1  # Convert to zero-based index

    initial_value = item_values[column_index]
    new_value = simpledialog.askstring("Edit", f"Edit {tree.heading(column_index)['text']}:", initialvalue=initial_value)

    if new_value:
        try:
            if column_index == 1:  # Amount column
                new_value = float(new_value)
            # Update the specific cell in the worksheet
            WS[int(item_id) + 1][column_index].value = new_value
            WB.save("input_data.xlsx")
            update_gui()
        except ValueError:
            messagebox.showerror("Error", "Invalid input. Please enter a valid value.")

    tree.bind("<Double-1>", on_treeview_double_click)

def delete_row():
    """Deletes a selected row from the Treeview and Excel sheet."""
    selected_item = tree.selection()[0]
    WS.delete_rows(int(selected_item) + 1)
    WB.save("input_data.xlsx")
    update_gui()

# Create the main window
root = tk.Tk()
root.title("Budget Tracker")
root.geometry("700x725")

# Apply the Forest theme
style = ttk.Style(root)
root.tk.call("source", r"C:\CodeProjects\Python_Budget\Forest-ttk-theme-master\forest-dark.tcl")
style.theme_use("forest-dark")

# Create a frame for input and buttons, and the chart button
input_chart_frame = ttk.Frame(root, padding="10")
input_chart_frame.grid(row=0, column=0, sticky="w", padx=10, pady=10)

# Create input labels and entry fields for expense
description_label = ttk.Label(input_chart_frame, text="Expense Description:")
description_label.grid(row=0, column=0, pady=5, padx=10, sticky="w")
description_entry = ttk.Entry(input_chart_frame, width=20)
description_entry.grid(row=0, column=1, pady=5, padx=10, sticky="w")

amount_label = ttk.Label(input_chart_frame, text="Money Spent:")
amount_label.grid(row=1, column=0, pady=5, padx=10, sticky="w")

# Function to validate amount entry
def validate_amount(value_if_allowed):
    return value_if_allowed.replace('.', '', 1).isdigit() or value_if_allowed == ""

validate_amount_cmd = root.register(validate_amount)
amount_entry = ttk.Entry(input_chart_frame, validate="key", validatecommand=(validate_amount_cmd, '%P'), width=20)
amount_entry.grid(row=1, column=1, pady=5, padx=10, sticky="w")

# Create a dropdown menu for categories
categories = ["Groceries", "Utilities", "Rent/Mortgage", "Entertainment", "Transportation", "Other", "Income"]
category_label = ttk.Label(input_chart_frame, text="Category:")
category_label.grid(row=2, column=0, pady=5, padx=10, sticky="w")
category_combo = ttk.Combobox(input_chart_frame, values=categories, width=17)
category_combo.grid(row=2, column=1, pady=5, padx=10, sticky="w")

# Create input labels and entry fields for income
income_description_label = ttk.Label(input_chart_frame, text="Income Description:")
income_description_label.grid(row=3, column=0, pady=5, padx=10, sticky="w")
income_description_entry = ttk.Entry(input_chart_frame, width=20)
income_description_entry.grid(row=3, column=1, pady=5, padx=10, sticky="w")

income_amount_label = ttk.Label(input_chart_frame, text="Income Amount:")
income_amount_label.grid(row=4, column=0, pady=5, padx=10, sticky="w")
income_amount_entry = ttk.Entry(input_chart_frame, width=20)
income_amount_entry.grid(row=4, column=1, pady=5, padx=10, sticky="w")

# Create a date entry for manual input
date_label = ttk.Label(input_chart_frame, text="Date (MM/DD/YYYY):")
date_label.grid(row=5, column=0, pady=5, padx=10, sticky="w")
date_entry = ttk.Entry(input_chart_frame, width=20)
date_entry.grid(row=5, column=1, pady=5, padx=10, sticky="w")
date_entry.insert(0, datetime.now().strftime("%m/%d/%Y"))

# Display total income
income_label_display = ttk.Label(input_chart_frame, text="Total Income: $0.00", font=("Calibre", 12, "bold"))
income_label_display.grid(row=0, column=3, pady=5, padx=10, sticky="w")

# Create a total expense label
total_label = ttk.Label(input_chart_frame, text="Total Expenses: $0.00", font=("Calibre", 12, "bold"))
total_label.grid(row=1, column=3, pady=5, padx=10, sticky="w")

# Create a balance label to show the positive balance
balance_label = ttk.Label(input_chart_frame, text="Balance: $0.00", font=("Calibre", 12, "bold"))
balance_label.grid(row=2, column=3, pady=5, padx=10, sticky="w")

# Create a frame for buttons
button_frame = ttk.Frame(input_chart_frame, padding="10")
button_frame.grid(row=6, column=0, columnspan=4, sticky="w")

# Create buttons for saving data and exporting data to CSV
save_button = ttk.Button(button_frame, text="Save to Excel", command=save_to_excel, style='Accent.TButton')
save_button.grid(row=0, column=0, padx=5, pady=5)

export_button = ttk.Button(button_frame, text="Export to CSV", command=export_to_csv, style='Accent.TButton')
export_button.grid(row=0, column=1, padx=5, pady=5)

# Add button to open charts window
chart_button = ttk.Button(button_frame, text="Show Charts", command=open_charts_window, style='Accent.TButton')
chart_button.grid(row=0, column=2, padx=5, pady=5)

# Add delete button to button frame
delete_button = ttk.Button(button_frame, text="Delete Row", command=delete_row, style='Accent.TButton')
delete_button.grid(row=0, column=3, padx=5, pady=5)

# Create a frame to hold the Treeview and the chart button
content_frame = ttk.Frame(root)
content_frame.grid(row=1, column=0, sticky="nw", padx=10, pady=5)

# Create a Treeview to display the data
tree_frame = ttk.Frame(content_frame)
tree_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=5)

tree = ttk.Treeview(tree_frame, columns=("Description", "Amount", "Category", "Date"), show="headings")
tree.heading("Description", text="Description")
tree.heading("Amount", text="Amount")
tree.heading("Category", text="Category")
tree.heading("Date", text="Date")

# Set column widths and alignment
tree.column("Description", width=150, anchor=tk.CENTER)
tree.column("Amount", width=100, anchor=tk.CENTER)
tree.column("Category", width=100, anchor=tk.CENTER)
tree.column("Date", width=100, anchor=tk.CENTER)

# Configure the Treeview to alternate row colors and add lines for better visibility
style.configure("Treeview.Heading", background="gray", foreground="white",
                font=("Calibre", 12, "bold"),relief="solid",)
tree.tag_configure("oddrow", background="gray25",font=("Calibre", 11))
tree.tag_configure("evenrow", background="gray30",font=("Calibre", 11))

tree.grid(row=0, column=0, sticky="nsew")

# Add a scrollbar
tree_scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
tree_scroll.grid(row=0, column=1, sticky="ns")
tree.configure(yscrollcommand=tree_scroll.set)

# Create a status label to show save status
status_label = ttk.Label(root, text="")
status_label.grid(row=2, column=0, pady=5, sticky="w", padx=10)

# Add padding around all widgets within the input and chart frame
for widget in input_chart_frame.winfo_children():
    widget.grid_configure(padx=5, pady=5)

# Bind Treeview for double-click to edit
tree.bind("<Double-1>", on_treeview_double_click)

# Load the data initially when the application starts
update_gui()

# Run the Tkinter event loop
root.mainloop()